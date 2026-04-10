from __future__ import annotations
import io
import calendar
from decimal import Decimal, InvalidOperation
from datetime import datetime
from app.config import now_cst
from fastapi import APIRouter, Request, Depends, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.database import get_db
from app.models.user import User, Profile
from app.models.finance import FinEmployee, FinSalaryRecord, FinExpense, FinFundRecord, FinMonthlyPayment, FinFundRequest, FinFundUsage, FinRewardPenalty
from app.models.team import Team
from app.auth import is_admin, is_finance

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


def _user(request: Request, db: Session) -> User | None:
    uid = request.session.get("user_id")
    if not uid:
        return None
    return db.query(User).filter(User.id == uid).first()


def _user_names(user: User, db: Session) -> set:
    """Return all name aliases for a user to match against applicant/tg_name fields.
    Includes login uname, profile.tg, and matching FinEmployee.tg_name (case-insensitive).
    """
    names = {user.uname}
    p = db.query(Profile).filter(Profile.uid == user.id).first()
    if p and p.tg:
        names.add(p.tg)
    # Find FinEmployee where tg_name matches any known name (case-insensitive)
    for name in list(names):
        emp = db.query(FinEmployee).filter(
            FinEmployee.tg_name.ilike(name)
        ).first()
        if emp and emp.tg_name:
            names.add(emp.tg_name)
    return names


def ok(data=None, message="success"):
    return {"success": True, "message": message, "data": data}


def err(message):
    return {"success": False, "message": message}


def to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        s = str(val).strip()
        if not s or s.lower() in ("none", "null", "-", ""):
            return None
        # 去掉非数字字符（保留小数点和负号）
        import re
        s = re.sub(r"[^\d.\-]", "", s)
        if not s:
            return None
        d = Decimal(s).quantize(Decimal("0.0001"))
        # 超过合理范围视为无效
        if abs(d) > Decimal("999999999"):
            return None
        return d
    except (InvalidOperation, Exception):
        return None


def now_str():
    return now_cst().strftime("%Y-%m-%d %H:%M:%S")


def _month_period(month_str: str) -> str:
    """Return period string like '3.1~3.31' for a YYYY-MM month string."""
    try:
        y, m = int(month_str[:4]), int(month_str[5:7])
        last = calendar.monthrange(y, m)[1]
        return f"{m}.1~{m}.{last}"
    except Exception:
        return ""


def _parse_base_salary(emp) -> Decimal | None:
    """Extract numeric base salary from employee adjusted_salary or formal_salary."""
    raw = (emp.adjusted_salary or emp.formal_salary) if emp else None
    return to_decimal(raw)


def _approved_expenses(db, month: str, tg_name: str) -> tuple:
    """Return (expense_rmb, expense_u) for APPROVED expenses of an employee in a given month."""
    rows = db.query(FinExpense).filter(
        FinExpense.expense_month == month,
        FinExpense.applicant == tg_name,
        FinExpense.status == "APPROVED",
    ).all()
    rmb = sum(float(r.amount) for r in rows if r.amount and r.currency in ("CNY", "RMB", None, ""))
    u = sum(float(r.amount) for r in rows if r.amount and r.currency == "U")
    return (
        Decimal(str(round(rmb, 4))) if rmb else Decimal("0"),
        Decimal(str(round(u, 4))) if u else Decimal("0"),
    )


def _parse_salary_num(val: str) -> str:
    """Strip unit suffixes from salary strings like '5000U', '22K', '2500元'.
    Returns a clean numeric string, or empty string if unparseable."""
    import re
    s = val.strip()
    # handle K/k suffix (e.g. 22K -> 22000)
    k_match = re.match(r'^([\d.]+)[Kk]$', s)
    if k_match:
        try:
            return str(int(float(k_match.group(1)) * 1000))
        except ValueError:
            pass
    # strip any trailing/leading non-numeric characters
    cleaned = re.sub(r'[^\d.]', '', s)
    return cleaned if cleaned else ""


# ─── 页面路由 ────────────────────────────────────────────────

@router.get("/finance/employees", response_class=HTMLResponse)
def page_employees(request: Request, keyword: str = "", db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    if keyword:
        rows = db.query(FinEmployee).filter(
            (FinEmployee.tg_name.contains(keyword)) | (FinEmployee.position.contains(keyword))
        ).all()
    else:
        rows = db.query(FinEmployee).order_by(FinEmployee.id.desc()).all()
    teams = db.query(Team).order_by(Team.name).all()
    # Build address map from Profile (single source of truth)
    profiles = db.query(Profile).all()
    addr_map = {}
    for p in profiles:
        key = (p.tg or "").strip().lower()
        if key:
            addr_map[key] = {"saddr": p.saddr or "", "vaddr": p.vaddr or ""}
    return templates.TemplateResponse("finance/employees.html", {
        "request": request, "employees": rows, "keyword": keyword,
        "teams": teams, "addr_map": addr_map, "is_admin": is_admin(user),
    })


@router.get("/finance/salary", response_class=HTMLResponse)
def page_salary(request: Request, year: str = "", month: str = "", db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)

    years = [r[0] for r in db.execute(text(
        "SELECT DISTINCT year_label FROM fin_salary_record WHERE year_label IS NOT NULL ORDER BY year_label DESC"
    )).fetchall()]
    months = [r[0] for r in db.execute(text(
        "SELECT DISTINCT month FROM fin_salary_record WHERE month IS NOT NULL ORDER BY month DESC"
    )).fetchall()]

    q = db.query(FinSalaryRecord)
    if year:
        q = q.filter(FinSalaryRecord.year_label == year)
    if month:
        q = q.filter(FinSalaryRecord.month == month)
    rows = q.order_by(FinSalaryRecord.month.desc(), FinSalaryRecord.id).all()

    from collections import OrderedDict
    # Group by month -> tg_name (no team sub-grouping needed for this view)
    grouped = OrderedDict()  # month -> list of plain dicts
    for r in rows:
        mk = r.month or "未知月份"
        if mk not in grouped:
            grouped[mk] = []
        grouped[mk].append({
            "id": r.id, "month": r.month, "tg_name": r.tg_name, "position": r.position,
            "base_salary": float(r.base_salary) if r.base_salary else None,
            "performance_salary": float(r.performance_salary) if r.performance_salary else None,
            "bonus": float(r.bonus) if r.bonus else None,
            "expense_rmb": float(r.expense_rmb) if r.expense_rmb else None,
            "expense_u": float(r.expense_u) if r.expense_u else None,
            "paid_rmb": float(r.paid_rmb) if r.paid_rmb else None,
            "period": r.period, "remarks": r.remarks,
            "wallet_address": r.wallet_address,
            "exchange_rate": float(r.exchange_rate) if r.exchange_rate else None,
            "paid_u": float(r.paid_u) if r.paid_u else None,
            "actual_amount": float(r.actual_amount) if r.actual_amount else None,
            "payment_date": r.payment_date, "status": r.status, "source": r.source,
        })

    return templates.TemplateResponse("finance/salary.html", {
        "request": request, "records": rows, "years": years, "months": months,
        "selectedYear": year, "selectedMonth": month, "grouped": grouped,
    })


@router.get("/finance/expenses", response_class=HTMLResponse)
def page_expenses(request: Request, status: str = "", db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if status:
        rows = db.query(FinExpense).filter(FinExpense.status == status).order_by(FinExpense.id.desc()).all()
    else:
        rows = db.query(FinExpense).order_by(FinExpense.id.desc()).all()
    # Only admin sees all records; everyone else sees only their own
    if not is_admin(user):
        user_names = _user_names(user, db)
        rows = [r for r in rows if r.applicant in user_names]
    emp_names = [e.tg_name for e in db.query(FinEmployee).order_by(FinEmployee.team, FinEmployee.tg_name).all() if e.tg_name]
    if not is_admin(user):
        user_names = _user_names(user, db)
        emp = db.query(FinEmployee).filter(FinEmployee.tg_name.in_(user_names)).first()
        display_name = emp.tg_name if emp else user.uname
    else:
        display_name = user.uname
    return templates.TemplateResponse("finance/expenses.html", {
        "request": request, "expenses": rows, "selectedStatus": status,
        "is_admin": is_admin(user), "uname": display_name, "emp_names": emp_names,
    })


@router.get("/finance/expenseslist", response_class=HTMLResponse)
def page_expenses_list(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    rows = db.query(FinExpense).order_by(FinExpense.expense_month.desc(), FinExpense.id.desc()).all()
    # group by expense_month, fall back to expense_date month
    from collections import OrderedDict
    grouped = OrderedDict()
    for r in rows:
        key = r.expense_month or (r.expense_date or "")[:7] or "未知月份"
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(r)
    return templates.TemplateResponse("finance/expenseslist.html", {
        "request": request, "grouped": grouped, "is_admin": is_admin(user),
    })


@router.get("/finance/funds", response_class=HTMLResponse)
def page_funds(request: Request, period: str = "", db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    periods = [r[0] for r in db.execute(text("SELECT DISTINCT period FROM fin_fund_record WHERE period IS NOT NULL ORDER BY id DESC")).fetchall()]
    if period:
        rows = db.query(FinFundRecord).filter(FinFundRecord.period == period).order_by(FinFundRecord.id.desc()).all()
    else:
        rows = db.query(FinFundRecord).order_by(FinFundRecord.id.desc()).all()
    return templates.TemplateResponse("finance/funds.html", {"request": request, "records": rows, "periods": periods, "selectedPeriod": period})


@router.get("/finance/reports", response_class=HTMLResponse)
def page_reports(request: Request, db: Session = Depends(get_db)):
    """综合报表: salary / expense / fund-request usage analysis."""
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)

    from sqlalchemy import func

    # ── 1. 工资报表: group by year_label, month ──
    salary_rows = db.execute(text("""
        SELECT year_label, month,
               SUM(CAST(paid_rmb AS DECIMAL(30,4))) AS total_rmb,
               SUM(CAST(paid_u   AS DECIMAL(30,4))) AS total_u,
               SUM(CAST(actual_amount AS DECIMAL(30,4))) AS total_actual,
               SUM(CAST(expense_rmb AS DECIMAL(30,4))) AS total_exp_rmb,
               SUM(CAST(expense_u   AS DECIMAL(30,4))) AS total_exp_u,
               COUNT(*) AS cnt
        FROM fin_salary_record
        GROUP BY year_label, month
        ORDER BY month DESC
    """)).fetchall()

    salary_by_year = {}
    for r in salary_rows:
        yr = r[0] or (r[1][:4] if r[1] else "?")
        if yr not in salary_by_year:
            salary_by_year[yr] = []
        salary_by_year[yr].append({
            "month": r[1], "total_rmb": float(r[2] or 0),
            "total_u": float(r[3] or 0), "total_actual": float(r[4] or 0),
            "total_exp_rmb": float(r[5] or 0), "total_exp_u": float(r[6] or 0),
            "cnt": r[7],
        })

    # ── 2. 报销报表: group by expense_month ──
    expense_rows = db.execute(text("""
        SELECT expense_month,
               SUM(CASE WHEN currency IN ('RMB','CNY','') OR currency IS NULL THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS rmb,
               SUM(CASE WHEN currency = 'U' THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS u_amt,
               COUNT(*) AS cnt,
               SUM(CASE WHEN status='APPROVED' THEN 1 ELSE 0 END) AS approved,
               SUM(CASE WHEN status='PENDING'  THEN 1 ELSE 0 END) AS pending,
               SUM(CASE WHEN status='REJECTED' THEN 1 ELSE 0 END) AS rejected
        FROM fin_expense
        GROUP BY expense_month
        ORDER BY expense_month DESC
    """)).fetchall()
    expense_data = [{"month": r[0], "rmb": float(r[1] or 0), "u": float(r[2] or 0),
                     "cnt": r[3], "approved": r[4], "pending": r[5], "rejected": r[6]}
                    for r in expense_rows]

    expense_by_year = {}
    for e in expense_data:
        yr = e["month"][:4] if e["month"] and len(e["month"]) >= 4 else "?"
        if yr not in expense_by_year:
            expense_by_year[yr] = []
        expense_by_year[yr].append(e)

    # ── 3. 资金申请 vs 使用: group by month ──
    req_rows = db.execute(text("""
        SELECT apply_month,
               SUM(CASE WHEN currency='U'   THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS req_u,
               SUM(CASE WHEN currency='RMB' THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS req_rmb,
               SUM(CASE WHEN status='APPROVED' AND currency='U'   THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS appr_u,
               SUM(CASE WHEN status='APPROVED' AND currency='RMB' THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS appr_rmb,
               COUNT(*) AS cnt
        FROM fin_fund_request
        GROUP BY apply_month ORDER BY apply_month DESC
    """)).fetchall()
    usage_rows = db.execute(text("""
        SELECT usage_month,
               SUM(CASE WHEN currency='U'   THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS use_u,
               SUM(CASE WHEN currency='RMB' THEN CAST(amount AS DECIMAL(30,4)) ELSE 0 END) AS use_rmb,
               COUNT(*) AS cnt
        FROM fin_fund_usage
        GROUP BY usage_month ORDER BY usage_month DESC
    """)).fetchall()

    req_map = {r[0]: {"req_u": float(r[1] or 0), "req_rmb": float(r[2] or 0),
                      "appr_u": float(r[3] or 0), "appr_rmb": float(r[4] or 0), "cnt": r[5]}
               for r in req_rows}
    use_map = {r[0]: {"use_u": float(r[1] or 0), "use_rmb": float(r[2] or 0), "cnt": r[3]}
               for r in usage_rows}
    all_months = sorted(set(list(req_map.keys()) + list(use_map.keys())), reverse=True)
    fund_flow = []
    for m in all_months:
        req = req_map.get(m, {"req_u": 0, "req_rmb": 0, "appr_u": 0, "appr_rmb": 0, "cnt": 0})
        use = use_map.get(m, {"use_u": 0, "use_rmb": 0, "cnt": 0})
        fund_flow.append({"month": m, **req, "use_u": use["use_u"], "use_rmb": use["use_rmb"],
                          "use_cnt": use["cnt"]})
    fund_by_year = {}
    for f in fund_flow:
        yr = f["month"][:4] if f["month"] and len(f["month"]) >= 4 else "?"
        if yr not in fund_by_year:
            fund_by_year[yr] = []
        fund_by_year[yr].append(f)

    years = sorted(set(list(salary_by_year.keys()) + list(expense_by_year.keys()) + list(fund_by_year.keys())), reverse=True)

    return templates.TemplateResponse("finance/reports.html", {
        "request": request,
        "salary_by_year": salary_by_year,
        "expense_by_year": expense_by_year,
        "fund_by_year": fund_by_year,
        "years": years,
    })


@router.get("/finance/import", response_class=HTMLResponse)
def page_import(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    return templates.TemplateResponse("finance/import.html", {"request": request})


# ─── REST: 员工档案 ────────────────────────────────────────


@router.get("/api/finance/employees")
def list_employees(db: Session = Depends(get_db)):
    return ok([_emp_dict(e) for e in db.query(FinEmployee).order_by(FinEmployee.id.desc()).all()])


@router.get("/api/finance/employees/{eid}")
def get_employee(eid: int, db: Session = Depends(get_db)):
    e = db.query(FinEmployee).filter(FinEmployee.id == eid).first()
    if not e:
        return err("不存在")
    return ok(_emp_dict(e))


@router.post("/api/finance/employees")
async def create_employee(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    e = FinEmployee(created_at=now_str(), updated_at=now_str(), **_emp_fields(body))
    db.add(e)
    db.commit()
    return ok()


@router.put("/api/finance/employees/{eid}")
async def update_employee(eid: int, request: Request, db: Session = Depends(get_db)):
    e = db.query(FinEmployee).filter(FinEmployee.id == eid).first()
    if not e:
        return err("不存在")
    body = await request.json()
    for k, v in _emp_fields(body).items():
        setattr(e, k, v)
    e.updated_at = now_str()
    db.commit()
    return ok()


@router.delete("/api/finance/employees/{eid}")
def delete_employee(eid: int, db: Session = Depends(get_db)):
    db.query(FinEmployee).filter(FinEmployee.id == eid).delete()
    db.commit()
    return ok()


@router.get("/api/finance/payinfo/{tg_name}")
def get_payinfo(tg_name: str, db: Session = Depends(get_db)):
    """Return QR image paths and wallet addresses for a given TG name."""
    # Try to find Profile by matching User.uname or Profile.tg (case-insensitive)
    profile = (
        db.query(Profile)
        .join(User, User.id == Profile.user_id)
        .filter(
            (User.uname.ilike(tg_name)) |
            (Profile.tg.ilike(tg_name))
        )
        .first()
    )
    if not profile:
        return ok({"saddr": None, "saddr_img": None, "vaddr": None, "vaddr_img": None})
    return ok({
        "saddr":     profile.saddr,
        "saddr_img": profile.saddr_img,
        "vaddr":     profile.vaddr,
        "vaddr_img": profile.vaddr_img,
    })


def _emp_fields(b):
    return {
        "tg_name": b.get("tgName") or b.get("tg_name"),
        "position": b.get("position"),
        "entry_date": b.get("entryDate") or b.get("entry_date"),
        "trial_salary": b.get("trialSalary") or b.get("trial_salary"),
        "formal_salary": b.get("formalSalary") or b.get("formal_salary"),
        "adjusted_salary": b.get("adjustedSalary") or b.get("adjusted_salary"),
        "salary_adjust_date": b.get("salaryAdjustDate") or b.get("salary_adjust_date"),
        "is_formal": b.get("isFormal") or b.get("is_formal"),
        "team": b.get("team"),
        "currency": b.get("currency"),
        "remarks": b.get("remarks"),
    }


def _emp_dict(e):
    return {"id": e.id, "tgName": e.tg_name, "position": e.position, "entryDate": e.entry_date,
            "trialSalary": e.trial_salary, "formalSalary": e.formal_salary, "adjustedSalary": e.adjusted_salary,
            "salaryAdjustDate": e.salary_adjust_date, "isFormal": e.is_formal, "team": e.team,
            "currency": e.currency, "remarks": e.remarks}


# ─── REST: 薪资记录 ────────────────────────────────────────

@router.get("/api/finance/salary")
def list_salary(db: Session = Depends(get_db)):
    return ok([_sal_dict(r) for r in db.query(FinSalaryRecord).order_by(FinSalaryRecord.id.desc()).all()])


@router.get("/api/finance/salary/{sid}")
def get_salary(sid: int, db: Session = Depends(get_db)):
    r = db.query(FinSalaryRecord).filter(FinSalaryRecord.id == sid).first()
    return ok(_sal_dict(r)) if r else err("不存在")


@router.post("/api/finance/salary")
async def create_salary(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    r = FinSalaryRecord(created_at=now_str(), updated_at=now_str(), **_sal_fields(body))
    db.add(r)
    db.commit()
    return ok()


@router.put("/api/finance/salary/{sid}")
async def update_salary(sid: int, request: Request, db: Session = Depends(get_db)):
    r = db.query(FinSalaryRecord).filter(FinSalaryRecord.id == sid).first()
    if not r:
        return err("不存在")
    body = await request.json()
    for k, v in _sal_fields(body).items():
        setattr(r, k, v)
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.delete("/api/finance/salary/{sid}")
def delete_salary(sid: int, db: Session = Depends(get_db)):
    db.query(FinSalaryRecord).filter(FinSalaryRecord.id == sid).delete()
    db.commit()
    return ok()


def _sal_fields(b):
    return {
        "tg_name": b.get("tgName"), "year_label": b.get("yearLabel"), "period": b.get("period"),
        "position": b.get("position"), "salary_type": b.get("salaryType"),
        "base_salary": to_decimal(b.get("baseSalary")),
        "performance_salary": to_decimal(b.get("performanceSalary")),
        "bonus": to_decimal(b.get("bonus")),
        "deduction": to_decimal(b.get("deduction")),
        "expense_rmb": to_decimal(b.get("expenseRmb")),
        "expense_u": to_decimal(b.get("expenseU")),
        "paid_rmb": to_decimal(b.get("paidRmb")), "paid_u": to_decimal(b.get("paidU")),
        "exchange_rate": to_decimal(b.get("exchangeRate")),
        "actual_amount": to_decimal(b.get("actualAmount")),
        "wallet_address": b.get("walletAddress"),
        "payment_date": b.get("paymentDate"),
        "remarks": b.get("remarks"),
    }


def _sal_dict(r):
    return {
        "id": r.id, "yearLabel": r.year_label, "month": r.month,
        "tgName": r.tg_name, "position": r.position, "period": r.period,
        "baseSalary": float(r.base_salary) if r.base_salary else None,
        "performanceSalary": float(r.performance_salary) if r.performance_salary else None,
        "bonus": float(r.bonus) if r.bonus else None,
        "paidRmb": float(r.paid_rmb) if r.paid_rmb else None,
        "remarks": r.remarks,
        "walletAddress": r.wallet_address,
        "exchangeRate": float(r.exchange_rate) if r.exchange_rate else None,
        "paidU": float(r.paid_u) if r.paid_u else None,
        "actualAmount": float(r.actual_amount) if r.actual_amount else None,
        "paymentDate": r.payment_date, "status": r.status, "source": r.source,
    }


# ─── REST: 费用报销 ────────────────────────────────────────

@router.get("/api/finance/expenses")
def list_expenses(db: Session = Depends(get_db)):
    return ok([_exp_dict(e) for e in db.query(FinExpense).order_by(FinExpense.id.desc()).all()])


@router.get("/api/finance/expenses/{eid}")
def get_expense(eid: int, db: Session = Depends(get_db)):
    e = db.query(FinExpense).filter(FinExpense.id == eid).first()
    return ok(_exp_dict(e)) if e else err("不存在")


@router.post("/api/finance/expenses")
async def create_expense(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    body = await request.json()
    fields = _exp_fields(body)
    # Non-admin can only submit for themselves — use canonical TG name
    if not user or not is_finance(user):
        user_names = _user_names(user, db) if user else set()
        emp = db.query(FinEmployee).filter(FinEmployee.tg_name.in_(user_names)).first()
        fields["applicant"] = emp.tg_name if emp else (user.uname if user else fields.get("applicant"))
    e = FinExpense(status="PENDING", created_at=now_str(), updated_at=now_str(), **fields)
    db.add(e)
    db.commit()
    return ok()


@router.post("/api/finance/expenses/batch")
async def batch_create_expenses(request: Request, db: Session = Depends(get_db)):
    """Create multiple expense records at once (from OCR multi-item detection)."""
    user = _user(request, db)
    body = await request.json()
    items = body.get("items", [])
    if not items:
        return err("items 不能为空")
    created = 0
    canonical_name = None
    if user and not is_finance(user):
        user_names = _user_names(user, db)
        emp = db.query(FinEmployee).filter(FinEmployee.tg_name.in_(user_names)).first()
        canonical_name = emp.tg_name if emp else user.uname
    for item in items:
        fields = _exp_fields(item)
        if not user or not is_finance(user):
            fields["applicant"] = canonical_name or fields.get("applicant")
        e = FinExpense(status="PENDING", created_at=now_str(), updated_at=now_str(), **fields)
        db.add(e)
        created += 1
    db.commit()
    return ok(message=f"已创建 {created} 条报销记录")


@router.put("/api/finance/expenses/{eid}")
async def update_expense(eid: int, request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    e = db.query(FinExpense).filter(FinExpense.id == eid).first()
    if not e:
        return err("不存在")
    body = await request.json()
    fields = _exp_fields(body)
    # Non-admin can only update their own and cannot change applicant
    if not user or not is_finance(user):
        user_names = _user_names(user, db) if user else set()
        if e.applicant and e.applicant not in user_names:
            return err("无权修改他人报销单")
        fields["applicant"] = e.applicant  # keep original canonical name
    for k, v in fields.items():
        setattr(e, k, v)
    e.updated_at = now_str()
    db.commit()
    return ok()


@router.put("/api/finance/expenses/{eid}/approve")
def approve_expense(eid: int, db: Session = Depends(get_db)):
    e = db.query(FinExpense).filter(FinExpense.id == eid).first()
    if not e:
        return err("不存在")
    e.status = "APPROVED"
    e.updated_at = now_str()
    db.commit()
    return ok()


@router.put("/api/finance/expenses/{eid}/reject")
def reject_expense(eid: int, db: Session = Depends(get_db)):
    e = db.query(FinExpense).filter(FinExpense.id == eid).first()
    if not e:
        return err("不存在")
    e.status = "REJECTED"
    e.updated_at = now_str()
    db.commit()
    return ok()


@router.delete("/api/finance/expenses/{eid}")
def delete_expense(eid: int, db: Session = Depends(get_db)):
    db.query(FinExpense).filter(FinExpense.id == eid).delete()
    db.commit()
    return ok()


def _exp_fields(b):
    return {"category": b.get("category"), "description": b.get("description"),
            "amount": to_decimal(b.get("amount")), "currency": b.get("currency"),
            "expense_date": b.get("expenseDate"), "applicant": b.get("applicant"),
            "image_path": b.get("imagePath"), "remarks": b.get("remarks"),
            "expense_month": b.get("expenseMonth")}


def _exp_dict(e):
    return {"id": e.id, "category": e.category, "description": e.description,
            "amount": str(e.amount) if e.amount else None, "currency": e.currency,
            "expenseDate": e.expense_date, "expenseMonth": e.expense_month,
            "applicant": e.applicant, "status": e.status,
            "imagePath": e.image_path, "remarks": e.remarks}


@router.post("/api/finance/expenses/ocr")
async def ocr_expense(file: UploadFile = File(...)):
    """Upload receipt image, extract amount via Gemini vision, return result."""
    import os, uuid
    from app.config import GEMINI_API_KEY

    content = await file.read()
    ext = (file.filename or "img.jpg").rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "gif", "webp"):
        return err("不支持的图片格式，请上传 jpg/png/webp")

    # Save image locally
    fname = f"{uuid.uuid4().hex}.{ext}"
    fpath = os.path.join("uploads", "expenses", fname)
    with open(fpath, "wb") as f:
        f.write(content)

    image_url = f"/uploads/expenses/{fname}"

    if not GEMINI_API_KEY:
        return ok(data={"amount": None, "currency": "CNY", "description": "", "imagePath": image_url},
                  message="图片已上传，未配置 GEMINI_API_KEY，请手动填写金额")

    try:
        import json, re
        from google import genai
        from google.genai import types

        client = genai.Client(api_key=GEMINI_API_KEY)
        media_type = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                types.Part.from_bytes(data=content, mime_type=media_type),
                "这是一张报销凭证图片，可能包含一条或多条报销项目。\n"
                "请识别所有报销项目，只返回JSON数组格式，不要任何解释：\n"
                "[{\"amount\": <数字或null>, \"currency\": \"CNY或USD或U\", \"description\": \"<简短描述>\"}]\n"
                "amount只返回纯数字（不含货币符号）。如果看不清或无法识别，对应字段返回null。\n"
                "如果图片中只有一条报销记录，数组只含一个元素。"
            ]
        )

        text = response.text.strip()
        m = re.search(r'\[.*\]', text, re.DOTALL)
        items = json.loads(m.group()) if m else []
        # Fallback: try single-object format
        if not items:
            m2 = re.search(r'\{.*\}', text, re.DOTALL)
            if m2:
                items = [json.loads(m2.group())]

        return ok(data={
            "items": [{"amount": it.get("amount"), "currency": it.get("currency", "CNY"),
                       "description": it.get("description", "")} for it in items],
            "imagePath": image_url,
        })
    except Exception as ex:
        return ok(data={"items": [], "imagePath": image_url},
                  message=f"图片已上传，识别失败: {ex}")


# ─── REST: 资金报表 ────────────────────────────────────────

@router.get("/api/finance/funds")
def list_funds(db: Session = Depends(get_db)):
    return ok([_fund_dict(r) for r in db.query(FinFundRecord).order_by(FinFundRecord.id.desc()).all()])


@router.get("/api/finance/funds/{fid}")
def get_fund(fid: int, db: Session = Depends(get_db)):
    r = db.query(FinFundRecord).filter(FinFundRecord.id == fid).first()
    return ok(_fund_dict(r)) if r else err("不存在")


@router.post("/api/finance/funds")
async def create_fund(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    r = FinFundRecord(created_at=now_str(), updated_at=now_str(), **_fund_fields(body))
    db.add(r)
    db.commit()
    return ok()


@router.put("/api/finance/funds/{fid}")
async def update_fund(fid: int, request: Request, db: Session = Depends(get_db)):
    r = db.query(FinFundRecord).filter(FinFundRecord.id == fid).first()
    if not r:
        return err("不存在")
    body = await request.json()
    for k, v in _fund_fields(body).items():
        setattr(r, k, v)
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.delete("/api/finance/funds/{fid}")
def delete_fund(fid: int, db: Session = Depends(get_db)):
    db.query(FinFundRecord).filter(FinFundRecord.id == fid).delete()
    db.commit()
    return ok()


def _fund_fields(b):
    return {"period": b.get("period"), "network": b.get("network"), "asset_type": b.get("assetType"),
            "recharge_amount": to_decimal(b.get("rechargeAmount")), "withdraw_amount": to_decimal(b.get("withdrawAmount")),
            "contract_balance": to_decimal(b.get("contractBalance")), "owner_balance": to_decimal(b.get("ownerBalance")),
            "record_date": b.get("recordDate"), "remarks": b.get("remarks")}


def _fund_dict(r):
    return {"id": r.id, "period": r.period, "network": r.network, "assetType": r.asset_type,
            "rechargeAmount": str(r.recharge_amount) if r.recharge_amount else None,
            "withdrawAmount": str(r.withdraw_amount) if r.withdraw_amount else None,
            "contractBalance": str(r.contract_balance) if r.contract_balance else None,
            "ownerBalance": str(r.owner_balance) if r.owner_balance else None,
            "recordDate": r.record_date, "remarks": r.remarks}


# ─── 资金申请 ─────────────────────────────────────────────

@router.get("/finance/fundrequest", response_class=HTMLResponse)
def page_fundrequest(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    rows = db.query(FinFundRequest).order_by(FinFundRequest.apply_month.desc(), FinFundRequest.id.desc()).all()
    from collections import OrderedDict
    grouped = OrderedDict()
    for r in rows:
        mk = r.apply_month or "未知月份"
        if mk not in grouped:
            grouped[mk] = []
        grouped[mk].append({
            "id": r.id, "apply_month": r.apply_month, "category": r.category,
            "currency": r.currency, "amount": float(r.amount) if r.amount else None,
            "exchange_rate": float(r.exchange_rate) if r.exchange_rate else None,
            "purpose": r.purpose, "applicant": r.applicant,
            "status": r.status, "approved_by": r.approved_by,
            "received_at": r.received_at, "remarks": r.remarks,
        })
    grouped = OrderedDict(sorted(grouped.items(), key=lambda x: x[0], reverse=True))
    emp_names = [e.tg_name for e in db.query(FinEmployee).order_by(FinEmployee.team, FinEmployee.tg_name).all() if e.tg_name]
    return templates.TemplateResponse("finance/fundrequest.html", {
        "request": request, "grouped": grouped, "emp_names": emp_names,
        "is_admin": is_admin(user), "uname": user.uname,
    })


@router.get("/api/finance/fundrequest/{rid}")
def get_fundrequest(rid: int, db: Session = Depends(get_db)):
    r = db.query(FinFundRequest).filter(FinFundRequest.id == rid).first()
    if not r:
        return err("不存在")
    return ok({"id": r.id, "applyMonth": r.apply_month, "category": r.category,
               "currency": r.currency, "amount": float(r.amount) if r.amount else None,
               "exchangeRate": float(r.exchange_rate) if r.exchange_rate else None,
               "purpose": r.purpose, "applicant": r.applicant,
               "status": r.status, "approvedBy": r.approved_by,
               "receivedAt": r.received_at, "remarks": r.remarks})


@router.post("/api/finance/fundrequest")
async def create_fundrequest(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return err("未登录")
    body = await request.json()
    # Non-admin: applicant locked to self (use canonical TG name if in employee records)
    if not is_finance(user):
        user_names = _user_names(user, db)
        emp = db.query(FinEmployee).filter(FinEmployee.tg_name.in_(user_names)).first()
        applicant = emp.tg_name if emp else user.uname
    else:
        applicant = body.get("applicant") or user.uname
    r = FinFundRequest(
        apply_month=body.get("applyMonth"),
        category=body.get("category", "办公室费用"),
        currency=body.get("currency", "U"),
        amount=to_decimal(body.get("amount")),
        exchange_rate=to_decimal(body.get("exchangeRate")),
        purpose=body.get("purpose"),
        applicant=applicant,
        status="PENDING",
        remarks=body.get("remarks"),
        created_at=now_str(), updated_at=now_str(),
    )
    db.add(r)
    db.commit()
    return ok()


@router.put("/api/finance/fundrequest/{rid}")
async def update_fundrequest(rid: int, request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return err("未登录")
    r = db.query(FinFundRequest).filter(FinFundRequest.id == rid).first()
    if not r:
        return err("不存在")
    if not is_finance(user) and r.applicant not in _user_names(user, db):
        return err("无权编辑他人申请")
    body = await request.json()
    r.apply_month = body.get("applyMonth", r.apply_month)
    r.category = body.get("category", r.category)
    r.currency = body.get("currency", r.currency)
    r.amount = to_decimal(body.get("amount")) if body.get("amount") is not None else r.amount
    r.exchange_rate = to_decimal(body.get("exchangeRate")) if body.get("exchangeRate") is not None else r.exchange_rate
    r.purpose = body.get("purpose", r.purpose)
    r.applicant = body.get("applicant", r.applicant) if is_finance(user) else r.applicant
    r.remarks = body.get("remarks", r.remarks)
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.put("/api/finance/fundrequest/{rid}/approve")
async def approve_fundrequest(rid: int, request: Request, db: Session = Depends(get_db)):
    r = db.query(FinFundRequest).filter(FinFundRequest.id == rid).first()
    if not r:
        return err("不存在")
    body = await request.json()
    r.status = "APPROVED"
    r.approved_by = body.get("approvedBy", "")
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.put("/api/finance/fundrequest/{rid}/reject")
def reject_fundrequest(rid: int, db: Session = Depends(get_db)):
    r = db.query(FinFundRequest).filter(FinFundRequest.id == rid).first()
    if not r:
        return err("不存在")
    r.status = "REJECTED"
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.put("/api/finance/fundrequest/{rid}/receive")
def receive_fundrequest(rid: int, request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return err("未登录")
    r = db.query(FinFundRequest).filter(FinFundRequest.id == rid).first()
    if not r:
        return err("不存在")
    if r.status != "APPROVED":
        return err("只有已批准的申请才能确认到账")
    # 申请人本人或管理员可确认
    if not is_finance(user) and r.applicant not in _user_names(user, db):
        return err("无权操作")
    r.received_at = now_str()
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.delete("/api/finance/fundrequest/{rid}")
def delete_fundrequest(rid: int, db: Session = Depends(get_db)):
    db.query(FinFundRequest).filter(FinFundRequest.id == rid).delete()
    db.commit()
    return ok()


# ─── 资金使用情况 ─────────────────────────────────────────

@router.get("/finance/fundusage", response_class=HTMLResponse)
def page_fundusage(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    from collections import OrderedDict, defaultdict

    # Approved fund requests grouped by month
    FR_CATEGORIES = ["办公室费用", "合约部署", "测试资金", "其他"]
    fr_rows = db.query(FinFundRequest).filter(
        FinFundRequest.status == "APPROVED",
        FinFundRequest.received_at.isnot(None),
    ).order_by(FinFundRequest.apply_month.desc()).all()
    fr_by_month = defaultdict(lambda: {"u": 0.0, "rmb": 0.0, "cny": 0.0, "records": [],
                                        "by_cat": {c: {"u": 0.0, "rmb": 0.0, "cny": 0.0} for c in FR_CATEGORIES}})
    for fr in fr_rows:
        m = fr.apply_month or "未知"
        amt = float(fr.amount) if fr.amount else 0.0
        rate = float(fr.exchange_rate) if fr.exchange_rate else 0.0
        cat = fr.category or "其他"
        if cat not in FR_CATEGORIES:
            cat = "其他"
        if fr.currency == "U":
            fr_by_month[m]["u"] += amt
            cny_equiv = amt * rate if rate else 0.0
            fr_by_month[m]["cny"] += cny_equiv
            fr_by_month[m]["by_cat"][cat]["u"] += amt
            fr_by_month[m]["by_cat"][cat]["cny"] += cny_equiv
        else:
            fr_by_month[m]["rmb"] += amt
            fr_by_month[m]["cny"] += amt
            fr_by_month[m]["by_cat"][cat]["rmb"] += amt
            fr_by_month[m]["by_cat"][cat]["cny"] += amt
        fr_by_month[m]["records"].append({
            "applicant": fr.applicant, "purpose": fr.purpose, "category": cat,
            "currency": fr.currency, "amount": amt, "exchange_rate": rate,
            "approved_by": fr.approved_by,
        })

    # Exchange rates per month from monthly payment records
    rate_rows = db.query(
        FinMonthlyPayment.month, FinMonthlyPayment.exchange_rate
    ).filter(FinMonthlyPayment.exchange_rate.isnot(None)).all()
    rate_by_month = {}
    for rr in rate_rows:
        if rr.month and rr.month not in rate_by_month and rr.exchange_rate:
            rate_by_month[rr.month] = float(rr.exchange_rate)

    # Usage records grouped by month, track U and CNY separately
    rows = db.query(FinFundUsage).order_by(FinFundUsage.usage_month.desc(), FinFundUsage.id.asc()).all()
    grouped = OrderedDict()
    usage_by_month = defaultdict(lambda: {
        "u": 0.0, "cny": 0.0,
        "by_cat": {c: {"u": 0.0, "cny": 0.0} for c in FR_CATEGORIES}
    })
    for r in rows:
        mk = r.usage_month or "未知月份"
        if mk not in grouped:
            grouped[mk] = []
        amt = float(r.amount) if r.amount else 0.0
        cat = r.category or "其他"
        if cat not in FR_CATEGORIES:
            cat = "其他"
        cur = r.currency or "CNY"
        grouped[mk].append({
            "id": r.id, "usage_month": r.usage_month, "currency": cur,
            "amount": float(r.amount) if r.amount else None,
            "category": cat, "description": r.description,
            "operator": r.operator, "remarks": r.remarks,
        })
        if cur == "U":
            usage_by_month[mk]["u"] += amt
            usage_by_month[mk]["by_cat"][cat]["u"] += amt
        else:
            usage_by_month[mk]["cny"] += amt
            usage_by_month[mk]["by_cat"][cat]["cny"] += amt

    # 已审批报销（APPROVED）也计入资金消耗
    approved_exps = db.query(FinExpense).filter(FinExpense.status == "APPROVED").all()
    total_exp_u = 0.0
    total_exp_cny = 0.0
    for e in approved_exps:
        amt = float(e.amount) if e.amount else 0.0
        mk = e.expense_month or "未知"
        cur = e.currency or "CNY"
        # 报销归入"办公室费用"类别
        cat = "办公室费用"
        if cur == "U":
            usage_by_month[mk]["u"] += amt
            usage_by_month[mk]["by_cat"][cat]["u"] += amt
            total_exp_u += amt
        else:
            usage_by_month[mk]["cny"] += amt
            usage_by_month[mk]["by_cat"][cat]["cny"] += amt
            total_exp_cny += amt

    total_fr_rmb = sum(v["rmb"] for v in fr_by_month.values())

    # 测试资金/合约部署 按U结算，办公室费用/其他 按CNY结算
    CAT_CURRENCY = {"办公室费用": "CNY", "合约部署": "U", "测试资金": "U", "其他": "CNY"}

    # Cumulative per-category totals
    cat_summary = {}
    for cat in FR_CATEGORIES:
        mode   = CAT_CURRENCY.get(cat, "CNY")
        fr_cny = sum(v["by_cat"].get(cat, {}).get("cny", 0.0) for v in fr_by_month.values())
        fr_u   = sum(v["by_cat"].get(cat, {}).get("u",   0.0) for v in fr_by_month.values())
        used_u   = sum(v["by_cat"].get(cat, {}).get("u",   0.0) for v in usage_by_month.values())
        used_cny = sum(v["by_cat"].get(cat, {}).get("cny", 0.0) for v in usage_by_month.values())
        if mode == "U":
            cat_summary[cat] = {
                "mode": "U", "fr": fr_u, "used": used_u, "remain": fr_u - used_u,
                "fr_u": fr_u, "fr_cny": fr_cny, "used_u": used_u, "used_cny": used_cny,
                "remain_u": fr_u - used_u, "remain_cny": 0.0,
            }
        else:
            cat_summary[cat] = {
                "mode": "CNY", "fr": fr_cny, "used": used_cny, "remain": fr_cny - used_cny,
                "fr_u": fr_u, "fr_cny": fr_cny, "used_u": used_u, "used_cny": used_cny,
                "remain_u": 0.0, "remain_cny": fr_cny - used_cny,
            }

    # Overall totals: U-mode categories count in U, CNY-mode categories count in CNY
    total_fr_u     = sum(cs["fr_u"]   for cat, cs in cat_summary.items() if CAT_CURRENCY.get(cat) == "U")
    total_fr_cny   = sum(cs["fr_cny"] for cat, cs in cat_summary.items() if CAT_CURRENCY.get(cat) == "CNY")
    total_used_u   = sum(cs["used_u"]   for cat, cs in cat_summary.items() if CAT_CURRENCY.get(cat) == "U")
    total_used_cny = sum(cs["used_cny"] for cat, cs in cat_summary.items() if CAT_CURRENCY.get(cat) == "CNY")
    balance_u   = total_fr_u   - total_used_u
    balance_cny = total_fr_cny - total_used_cny

    # Fund request list (for reference display)
    fr_list = []
    for fr in fr_rows:
        amt  = float(fr.amount) if fr.amount else 0.0
        rate = float(fr.exchange_rate) if fr.exchange_rate else 0.0
        cny  = amt * rate if fr.currency == "U" and rate else (amt if fr.currency != "U" else 0.0)
        fr_list.append({
            "apply_month": fr.apply_month, "category": fr.category or "其他",
            "currency": fr.currency, "amount": amt, "exchange_rate": rate,
            "cny": cny, "purpose": fr.purpose, "applicant": fr.applicant,
        })

    # All months (usage months only for monthly cards)
    all_months = sorted(grouped.keys(), reverse=True)

    return templates.TemplateResponse("finance/fundusage.html", {
        "request": request, "grouped": grouped,
        "usage_by_month": usage_by_month, "all_months": all_months,
        "total_fr_cny": total_fr_cny, "total_fr_u": total_fr_u, "total_fr_rmb": total_fr_rmb,
        "total_used_u": total_used_u, "total_used_cny": total_used_cny,
        "balance_u": balance_u, "balance_cny": balance_cny,
        "cat_summary": cat_summary, "fr_list": fr_list,
        "fr_categories": FR_CATEGORIES,
        "is_admin": is_admin(user), "my_view": False, "uname": user.uname,
        "total_exp_u": total_exp_u, "total_exp_cny": total_exp_cny,
    })


@router.get("/my/fundrequest", response_class=HTMLResponse)
def page_my_fundrequest(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login")
    user_names = _user_names(user, db)
    if is_admin(user):
        rows = db.query(FinFundRequest).order_by(FinFundRequest.apply_month.desc(), FinFundRequest.id.desc()).all()
    else:
        rows = db.query(FinFundRequest).filter(
            FinFundRequest.applicant.in_(user_names)
        ).order_by(FinFundRequest.id.desc()).all()
    from collections import OrderedDict
    grouped = OrderedDict()
    for r in rows:
        mk = r.apply_month or "未知月份"
        if mk not in grouped:
            grouped[mk] = []
        grouped[mk].append({
            "id": r.id, "apply_month": r.apply_month, "category": r.category,
            "currency": r.currency, "amount": float(r.amount) if r.amount else None,
            "exchange_rate": float(r.exchange_rate) if r.exchange_rate else None,
            "purpose": r.purpose, "applicant": r.applicant,
            "status": r.status, "approved_by": r.approved_by,
            "received_at": r.received_at, "remarks": r.remarks,
        })
    grouped = OrderedDict(sorted(grouped.items(), key=lambda x: x[0], reverse=True))
    emp_names = [e.tg_name for e in db.query(FinEmployee).order_by(FinEmployee.team, FinEmployee.tg_name).all() if e.tg_name]
    emp = db.query(FinEmployee).filter(FinEmployee.tg_name.in_(user_names)).first()
    display_name = emp.tg_name if emp else user.uname
    return templates.TemplateResponse("finance/fundrequest.html", {
        "request": request, "grouped": grouped, "my_view": True, "uname": display_name,
        "emp_names": emp_names, "is_admin": False,
    })


@router.get("/my/fundusage", response_class=HTMLResponse)
def page_my_fundusage(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login")
    from collections import OrderedDict, defaultdict
    MY_FR_CATS = ["办公室费用", "合约部署", "测试资金", "其他"]
    user_names = _user_names(user, db)
    if is_admin(user):
        rows = db.query(FinFundUsage).order_by(FinFundUsage.usage_month.desc(), FinFundUsage.id.asc()).all()
    else:
        rows = db.query(FinFundUsage).filter(
            FinFundUsage.operator.in_(user_names)
        ).order_by(FinFundUsage.usage_month.desc(), FinFundUsage.id.asc()).all()
    grouped = OrderedDict()
    usage_by_month = defaultdict(lambda: {
        "u": 0.0, "cny": 0.0,
        "by_cat": {c: {"u": 0.0, "cny": 0.0} for c in MY_FR_CATS}
    })
    for r in rows:
        mk = r.usage_month or "未知月份"
        if mk not in grouped:
            grouped[mk] = []
        amt = float(r.amount) if r.amount else 0.0
        cat = r.category or "其他"
        if cat not in MY_FR_CATS:
            cat = "其他"
        cur = r.currency or "CNY"
        grouped[mk].append({
            "id": r.id, "usage_month": r.usage_month, "currency": cur,
            "amount": float(r.amount) if r.amount else None,
            "category": cat, "description": r.description,
            "operator": r.operator, "remarks": r.remarks,
        })
        if cur == "U":
            usage_by_month[mk]["u"] += amt
            usage_by_month[mk]["by_cat"][cat]["u"] += amt
        else:
            usage_by_month[mk]["cny"] += amt
            usage_by_month[mk]["by_cat"][cat]["cny"] += amt
    MY_CAT_CURRENCY = {"办公室费用": "CNY", "合约部署": "U", "测试资金": "U", "其他": "CNY"}
    total_used_u   = sum(v["u"]   for v in usage_by_month.values())
    total_used_cny = sum(v["cny"] for v in usage_by_month.values())
    my_cat_summary = {}
    for cat in MY_FR_CATS:
        mode     = MY_CAT_CURRENCY.get(cat, "CNY")
        used_u   = sum(v["by_cat"].get(cat, {}).get("u",   0.0) for v in usage_by_month.values())
        used_cny = sum(v["by_cat"].get(cat, {}).get("cny", 0.0) for v in usage_by_month.values())
        used = used_u if mode == "U" else used_cny
        my_cat_summary[cat] = {
            "mode": mode, "fr": 0, "used": used, "remain": -used,
            "fr_u": 0, "fr_cny": 0, "used_u": used_u, "used_cny": used_cny,
            "remain_u": -used_u, "remain_cny": -used_cny,
        }
    all_months = sorted(grouped.keys(), reverse=True)
    return templates.TemplateResponse("finance/fundusage.html", {
        "request": request, "grouped": grouped,
        "usage_by_month": usage_by_month, "all_months": all_months,
        "total_fr_cny": 0, "total_fr_u": 0, "total_fr_rmb": 0,
        "total_used_u": total_used_u, "total_used_cny": total_used_cny,
        "balance_u": -total_used_u, "balance_cny": -total_used_cny,
        "cat_summary": my_cat_summary,
        "fr_list": [], "fr_categories": MY_FR_CATS,
        "is_admin": False, "uname": user.uname, "my_view": True,
    })


@router.get("/api/finance/fundusage/{uid}")
def get_fundusage(uid: int, db: Session = Depends(get_db)):
    r = db.query(FinFundUsage).filter(FinFundUsage.id == uid).first()
    if not r:
        return err("不存在")
    return ok({"id": r.id, "usageMonth": r.usage_month, "currency": r.currency,
               "amount": float(r.amount) if r.amount else None,
               "category": r.category, "description": r.description,
               "operator": r.operator, "remarks": r.remarks})


@router.post("/api/finance/fundusage")
async def create_fundusage(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return err("未登录")
    body = await request.json()
    # Operator: admin can specify freely; others auto-fill from session
    if is_finance(user):
        operator = body.get("operator") or user.uname
    else:
        operator = user.uname
    r = FinFundUsage(
        usage_month=body.get("usageMonth"),
        currency=body.get("currency", "U"),
        amount=to_decimal(body.get("amount")),
        category=body.get("category", "其他"),
        description=body.get("description"),
        operator=operator,
        remarks=body.get("remarks"),
        created_at=now_str(), updated_at=now_str(),
    )
    db.add(r)
    db.commit()
    return ok()


@router.put("/api/finance/fundusage/{uid}")
async def update_fundusage(uid: int, request: Request, db: Session = Depends(get_db)):
    r = db.query(FinFundUsage).filter(FinFundUsage.id == uid).first()
    if not r:
        return err("不存在")
    body = await request.json()
    r.usage_month = body.get("usageMonth", r.usage_month)
    r.currency = body.get("currency", r.currency)
    r.amount = to_decimal(body.get("amount")) if body.get("amount") is not None else r.amount
    r.category = body.get("category", r.category)
    r.description = body.get("description", r.description)
    r.operator = body.get("operator", r.operator)
    r.remarks = body.get("remarks", r.remarks)
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.delete("/api/finance/fundusage/{uid}")
def delete_fundusage(uid: int, db: Session = Depends(get_db)):
    db.query(FinFundUsage).filter(FinFundUsage.id == uid).delete()
    db.commit()
    return ok()


@router.post("/api/finance/fundusage/batch")
async def batch_create_fundusage(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    items = body.get("items", [])
    if not items:
        return err("无数据")
    created = 0
    for it in items:
        r = FinFundUsage(
            usage_month=it.get("usageMonth"),
            currency=it.get("currency", "U"),
            amount=to_decimal(it.get("amount")),
            category=it.get("category", "其他"),
            description=it.get("description"),
            operator=it.get("operator"),
            remarks=it.get("remarks"),
            created_at=now_str(), updated_at=now_str(),
        )
        db.add(r)
        created += 1
    db.commit()
    return ok(message=f"已创建 {created} 条记录")


@router.post("/api/finance/fundusage/ocr")
async def ocr_fundusage(file: UploadFile = File(...)):
    """Upload image, extract fund usage items via Gemini vision."""
    import os, uuid
    from app.config import GEMINI_API_KEY

    content = await file.read()
    ext = (file.filename or "img.jpg").rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "gif", "webp"):
        return err("不支持的图片格式，请上传 jpg/png/webp")

    fname = f"{uuid.uuid4().hex}.{ext}"
    fpath = os.path.join("uploads", "fundusage", fname)
    os.makedirs(os.path.dirname(fpath), exist_ok=True)
    with open(fpath, "wb") as f:
        f.write(content)
    image_url = f"/uploads/fundusage/{fname}"

    if not GEMINI_API_KEY:
        return ok(data={"items": [], "imagePath": image_url},
                  message="图片已上传，未配置 GEMINI_API_KEY，请手动填写")

    try:
        import json, re
        from google import genai
        from google.genai import types

        client = genai.Client(api_key=GEMINI_API_KEY)
        media_type = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                types.Part.from_bytes(data=content, mime_type=media_type),
                "这是一张资金使用凭证图片，可能包含一条或多条支出记录。\n"
                "请识别所有支出项目，只返回JSON数组格式，不要任何解释：\n"
                "[{\"amount\": <数字或null>, \"currency\": \"U或RMB\", "
                "\"category\": \"薪资/运营/推广/报销/其他\", \"description\": \"<简短描述>\"}]\n"
                "amount只返回纯数字（不含货币符号）。如果看不清，对应字段返回null。\n"
                "如果只有一条记录，数组只含一个元素。"
            ]
        )

        text = response.text.strip()
        m = re.search(r'\[.*\]', text, re.DOTALL)
        items = json.loads(m.group()) if m else []
        if not items:
            m2 = re.search(r'\{.*\}', text, re.DOTALL)
            if m2:
                items = [json.loads(m2.group())]

        return ok(data={
            "items": [{"amount": it.get("amount"), "currency": it.get("currency", "U"),
                       "category": it.get("category", "其他"),
                       "description": it.get("description", "")} for it in items],
            "imagePath": image_url,
        })
    except Exception as ex:
        return ok(data={"items": [], "imagePath": image_url},
                  message=f"图片已上传，识别失败: {ex}")


# ─── 员工奖惩 ─────────────────────────────────────────────

def _rp_net(db, month: str, tg_name: str, currency: str) -> tuple:
    """Return (reward, penalty) amounts in the given currency for an employee in a month."""
    rows = db.query(FinRewardPenalty).filter(
        FinRewardPenalty.month == month,
        FinRewardPenalty.tg_name == tg_name,
        FinRewardPenalty.currency == currency,
    ).all()
    reward  = sum(float(r.amount) for r in rows if r.type == "REWARD"  and r.amount)
    penalty = sum(float(r.amount) for r in rows if r.type == "PENALTY" and r.amount)
    return (
        Decimal(str(round(reward,  4))) if reward  else Decimal("0"),
        Decimal(str(round(penalty, 4))) if penalty else Decimal("0"),
    )


@router.get("/finance/rewardpenalty", response_class=HTMLResponse)
def page_rewardpenalty(request: Request, month: str = "", db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    from collections import OrderedDict
    q = db.query(FinRewardPenalty)
    if month:
        q = q.filter(FinRewardPenalty.month == month)
    rows = q.order_by(FinRewardPenalty.month.desc(), FinRewardPenalty.id.asc()).all()
    months_q = db.execute(text(
        "SELECT DISTINCT month FROM fin_reward_penalty WHERE month IS NOT NULL ORDER BY month DESC"
    )).fetchall()
    all_months = [r[0] for r in months_q]
    grouped = OrderedDict()
    for r in rows:
        mk = r.month or "未知月份"
        if mk not in grouped:
            grouped[mk] = []
        grouped[mk].append({
            "id": r.id, "month": r.month, "tg_name": r.tg_name,
            "type": r.type, "amount": float(r.amount) if r.amount else None,
            "currency": r.currency, "reason": r.reason,
        })
    emp_names = [e.tg_name for e in db.query(FinEmployee).order_by(FinEmployee.team, FinEmployee.tg_name).all() if e.tg_name]
    return templates.TemplateResponse("finance/rewardpenalty.html", {
        "request": request, "grouped": grouped, "emp_names": emp_names,
        "all_months": all_months, "selected_month": month, "uname": user.uname,
    })


@router.get("/api/finance/rewardpenalty/{rid}")
def get_rewardpenalty(rid: int, db: Session = Depends(get_db)):
    r = db.query(FinRewardPenalty).filter(FinRewardPenalty.id == rid).first()
    if not r:
        return err("不存在")
    return ok({"id": r.id, "month": r.month, "tgName": r.tg_name, "type": r.type,
               "amount": float(r.amount) if r.amount else None,
               "currency": r.currency, "reason": r.reason})


@router.post("/api/finance/rewardpenalty")
async def create_rewardpenalty(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user or not is_finance(user):
        return err("无权操作")
    body = await request.json()
    r = FinRewardPenalty(
        month=body.get("month"),
        tg_name=body.get("tgName"),
        type=body.get("type", "REWARD"),
        amount=to_decimal(body.get("amount")),
        currency=body.get("currency", "U"),
        reason=body.get("reason"),
        created_by=user.uname,
        created_at=now_str(), updated_at=now_str(),
    )
    db.add(r)
    db.commit()
    return ok()


@router.put("/api/finance/rewardpenalty/{rid}")
async def update_rewardpenalty(rid: int, request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user or not is_finance(user):
        return err("无权操作")
    r = db.query(FinRewardPenalty).filter(FinRewardPenalty.id == rid).first()
    if not r:
        return err("不存在")
    body = await request.json()
    r.month    = body.get("month", r.month)
    r.tg_name  = body.get("tgName", r.tg_name)
    r.type     = body.get("type", r.type)
    r.amount   = to_decimal(body.get("amount")) if body.get("amount") is not None else r.amount
    r.currency = body.get("currency", r.currency)
    r.reason   = body.get("reason", r.reason)
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.delete("/api/finance/rewardpenalty/{rid}")
def delete_rewardpenalty(rid: int, db: Session = Depends(get_db)):
    db.query(FinRewardPenalty).filter(FinRewardPenalty.id == rid).delete()
    db.commit()
    return ok()


# ─── 月度薪资发放 ──────────────────────────────────────────

DEFAULT_TEAMS = ["web3团队", "AI 团队", "分布式团队"]


@router.get("/finance/monthlypay", response_class=HTMLResponse)
def page_monthlypay(request: Request, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        return RedirectResponse("/login")
    if not is_finance(user):
        return HTMLResponse("无权访问", status_code=403)
    rows = db.query(FinMonthlyPayment).order_by(
        FinMonthlyPayment.month.desc(), FinMonthlyPayment.id.asc()
    ).all()

    # Build expense map: applicant -> {rmb, u} — APPROVED only
    from collections import OrderedDict, defaultdict
    exp_rows = db.query(FinExpense).filter(FinExpense.status == "APPROVED").all()
    exp_map = defaultdict(lambda: {"rmb": 0.0, "u": 0.0})
    for e in exp_rows:
        if not e.applicant:
            continue
        key = e.applicant
        amt = float(e.amount) if e.amount else 0.0
        if e.currency in ("CNY", "RMB", None, ""):
            exp_map[key]["rmb"] += amt
        elif e.currency == "U":
            exp_map[key]["u"] += amt

    # Build reward/penalty map: (month, tg_name) -> {reward_u, penalty_u, reward_rmb, penalty_rmb}
    rp_rows = db.query(FinRewardPenalty).all()
    rp_map = defaultdict(lambda: {"reward_u": 0.0, "penalty_u": 0.0, "reward_rmb": 0.0, "penalty_rmb": 0.0})
    for rp in rp_rows:
        if not rp.month or not rp.tg_name:
            continue
        key = (rp.month, rp.tg_name.lower())
        amt = float(rp.amount) if rp.amount else 0.0
        cur = rp.currency or "U"
        if rp.type == "REWARD":
            if cur == "U":
                rp_map[key]["reward_u"] += amt
            else:
                rp_map[key]["reward_rmb"] += amt
        else:
            if cur == "U":
                rp_map[key]["penalty_u"] += amt
            else:
                rp_map[key]["penalty_rmb"] += amt

    # Build approved fund requests map: (apply_month, applicant_lower) -> {u, rmb}
    fr_rows = db.query(FinFundRequest).filter(FinFundRequest.status == "APPROVED").all()
    fr_map = defaultdict(lambda: {"u": 0.0, "rmb": 0.0})
    for fr in fr_rows:
        m = fr.apply_month or ""
        name = (fr.applicant or "").lower()
        amt = float(fr.amount) if fr.amount else 0.0
        if fr.currency == "U":
            fr_map[(m, name)]["u"] += amt
        else:
            fr_map[(m, name)]["rmb"] += amt

    # group: month -> team -> [dicts]
    grouped = OrderedDict()
    for r in rows:
        m = r.month or "未知"
        t = r.team or "未分配"
        if m not in grouped:
            grouped[m] = OrderedDict()
        if t not in grouped[m]:
            grouped[m][t] = []
        exp = exp_map.get(r.tg_name, {"rmb": 0.0, "u": 0.0})
        rp = rp_map.get((r.month, (r.tg_name or "").lower()), {"reward_u": 0.0, "penalty_u": 0.0, "reward_rmb": 0.0, "penalty_rmb": 0.0})
        fr = fr_map.get((r.month or "", (r.tg_name or "").lower()), {"u": 0.0, "rmb": 0.0})
        grouped[m][t].append({
            "id": r.id,
            "month": r.month,
            "team": r.team,
            "tg_name": r.tg_name,
            "currency": r.currency,
            "amount": float(r.amount) if r.amount else None,
            "exchange_rate": float(r.exchange_rate) if r.exchange_rate else None,
            "wallet_address": r.wallet_address or "",
            "remarks": r.remarks,
            "status": r.status,
            "expense_rmb": exp["rmb"],
            "expense_u": exp["u"],
            "rp_reward_u": rp["reward_u"],
            "rp_penalty_u": rp["penalty_u"],
            "rp_reward_rmb": rp["reward_rmb"],
            "rp_penalty_rmb": rp["penalty_rmb"],
            "fr_u": fr["u"],
            "fr_rmb": fr["rmb"],
        })

    # Re-sort grouped by month descending
    grouped = OrderedDict(sorted(grouped.items(), reverse=True))

    return templates.TemplateResponse("finance/monthlypay.html", {
        "request": request, "grouped": grouped,
    })


@router.get("/api/finance/monthlypay/employees")
def monthlypay_employees(db: Session = Depends(get_db)):
    """Return employees grouped by team for pre-populating the batch create form."""
    emps = db.query(FinEmployee).filter(
        FinEmployee.team.in_(DEFAULT_TEAMS)
    ).order_by(FinEmployee.team, FinEmployee.id).all()

    # build tg -> profile address map (case-insensitive)
    profiles = db.query(Profile).all()
    addr_map = {(p.tg or "").lower(): {"saddr": p.saddr or "", "vaddr": p.vaddr or ""}
                for p in profiles if p.tg}

    result = {}
    for e in emps:
        team = e.team
        if team not in result:
            result[team] = []
        raw = e.adjusted_salary or e.formal_salary or e.trial_salary or ""
        salary = _parse_salary_num(str(raw)) if raw else ""
        prof_addrs = addr_map.get((e.tg_name or "").lower(), {})
        saddr = prof_addrs.get("saddr") or ""
        vaddr = prof_addrs.get("vaddr") or ""
        result[team].append({
            "tgName": e.tg_name,
            "salary": salary,
            "currency": e.currency or "",
            "saddr": saddr,
            "vaddr": vaddr,
        })
    for t in DEFAULT_TEAMS:
        if t not in result:
            result[t] = []
    return ok(result)


@router.post("/api/finance/monthlypay/batch")
async def create_monthlypay_batch(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    items = body.get("items", [])
    if not items:
        return err("无有效数据")
    for item in items:
        r = FinMonthlyPayment(
            month=item.get("month"),
            team=item.get("team"),
            tg_name=item.get("tgName"),
            currency=item.get("currency"),
            amount=to_decimal(item.get("amount")),
            exchange_rate=to_decimal(item.get("exchangeRate")) if item.get("exchangeRate") else None,
            wallet_address=item.get("walletAddress") or None,
            remarks=item.get("remarks"),
            status="UNPAID",
            created_at=now_str(),
            updated_at=now_str(),
        )
        db.add(r)
    db.commit()
    return ok()


@router.put("/api/finance/monthlypay/{rid}")
async def update_monthlypay(rid: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    r = db.query(FinMonthlyPayment).filter(FinMonthlyPayment.id == rid).first()
    if not r:
        return err("不存在")
    r.month = body.get("month", r.month)
    r.team = body.get("team", r.team)
    r.tg_name = body.get("tgName", r.tg_name)
    r.currency = body.get("currency", r.currency)
    r.amount = to_decimal(body.get("amount")) if body.get("amount") is not None else r.amount
    r.exchange_rate = to_decimal(body.get("exchangeRate")) if body.get("exchangeRate") else r.exchange_rate
    if "walletAddress" in body:
        r.wallet_address = body.get("walletAddress") or None
    r.remarks = body.get("remarks", r.remarks)
    r.updated_at = now_str()
    db.commit()
    return ok()


@router.put("/api/finance/monthlypay/{rid}/paid")
def mark_monthlypay_paid(rid: int, db: Session = Depends(get_db)):
    r = db.query(FinMonthlyPayment).filter(FinMonthlyPayment.id == rid).first()
    if not r:
        return err("不存在")
    r.status = "PAID"
    r.updated_at = now_str()

    # Sync to fin_salary_record — skip if already exists from this monthly pay record
    existing = db.query(FinSalaryRecord).filter(
        FinSalaryRecord.source == "monthly_pay",
        FinSalaryRecord.month == r.month,
        FinSalaryRecord.tg_name == r.tg_name,
    ).first()
    if not existing:
        year_label = r.month[:4] if r.month and len(r.month) >= 4 else None
        emp = db.query(FinEmployee).filter(FinEmployee.tg_name == r.tg_name).first()
        if not r.wallet_address:
            prof = (db.query(Profile).join(User, User.id == Profile.uid)
                    .filter((User.uname.ilike(r.tg_name)) | (Profile.tg.ilike(r.tg_name))).first())
            wallet = (prof.saddr or prof.vaddr) if prof else None
        else:
            wallet = r.wallet_address
        position = emp.position if emp else None
        base_salary = _parse_base_salary(emp)
        paid_u = None
        paid_rmb = None
        if r.currency == "U":
            paid_u = float(r.amount) if r.amount else None
            paid_rmb = float(r.amount * r.exchange_rate) if r.amount and r.exchange_rate else None
        else:
            paid_rmb = float(r.amount) if r.amount else None
            paid_u = float(r.amount / r.exchange_rate) if r.amount and r.exchange_rate else None
        exp_rmb, exp_u = _approved_expenses(db, r.month, r.tg_name)
        reward_amt, penalty_amt = _rp_net(db, r.month, r.tg_name, r.currency)
        paid_u_dec = to_decimal(paid_u)
        rate = float(r.exchange_rate) if r.exchange_rate else 0
        exp_rmb_u = float(exp_rmb or 0) / rate if rate else 0
        rp_net = float(reward_amt or 0) - float(penalty_amt or 0)
        total_u = float(paid_u_dec or 0) + float(exp_u or 0) + exp_rmb_u + (rp_net if r.currency == "U" else rp_net / rate if rate else 0)
        sr = FinSalaryRecord(
            year_label=year_label,
            month=r.month,
            tg_name=r.tg_name,
            position=position,
            base_salary=base_salary,
            performance_salary=Decimal("0"),
            bonus=Decimal("0"),
            expense_rmb=exp_rmb if exp_rmb else None,
            expense_u=exp_u if exp_u else None,
            reward_amount=reward_amt if reward_amt else None,
            penalty_amount=penalty_amt if penalty_amt else None,
            paid_rmb=to_decimal(paid_rmb),
            paid_u=paid_u_dec,
            actual_amount=to_decimal(total_u) if total_u else None,
            exchange_rate=r.exchange_rate,
            wallet_address=wallet,
            period=_month_period(r.month) if r.month else None,
            remarks=r.remarks,
            payment_date=now_str()[:10],
            status="PAID",
            source="monthly_pay",
            created_at=now_str(), updated_at=now_str(),
        )
        db.add(sr)

    db.commit()
    return ok()


@router.put("/api/finance/monthlypay/batch-paid/{month}")
def batch_mark_paid(month: str, db: Session = Depends(get_db)):
    """Mark all UNPAID records for a month as PAID and sync to fin_salary_record."""
    records = db.query(FinMonthlyPayment).filter(
        FinMonthlyPayment.month == month,
        FinMonthlyPayment.status != "PAID",
    ).all()
    if not records:
        return err("该月无待付记录")

    synced = 0
    for r in records:
        r.status = "PAID"
        r.updated_at = now_str()

        existing = db.query(FinSalaryRecord).filter(
            FinSalaryRecord.source == "monthly_pay",
            FinSalaryRecord.month == r.month,
            FinSalaryRecord.tg_name == r.tg_name,
        ).first()
        if existing:
            continue

        year_label = r.month[:4] if r.month and len(r.month) >= 4 else None
        emp = db.query(FinEmployee).filter(FinEmployee.tg_name == r.tg_name).first()
        if not r.wallet_address:
            prof = (db.query(Profile).join(User, User.id == Profile.uid)
                    .filter((User.uname.ilike(r.tg_name)) | (Profile.tg.ilike(r.tg_name))).first())
            wallet = (prof.saddr or prof.vaddr) if prof else None
        else:
            wallet = r.wallet_address
        position = emp.position if emp else None
        base_salary = _parse_base_salary(emp)

        if r.currency == "U":
            paid_u = float(r.amount) if r.amount else None
            paid_rmb = round(float(r.amount * r.exchange_rate), 2) if r.amount and r.exchange_rate else None
        else:
            paid_rmb = float(r.amount) if r.amount else None
            paid_u = round(float(r.amount / r.exchange_rate), 4) if r.amount and r.exchange_rate else None

        exp_rmb, exp_u = _approved_expenses(db, r.month, r.tg_name)
        reward_amt, penalty_amt = _rp_net(db, r.month, r.tg_name, r.currency)
        paid_u_dec = to_decimal(paid_u)
        rate = float(r.exchange_rate) if r.exchange_rate else 0
        exp_rmb_u = float(exp_rmb or 0) / rate if rate else 0
        rp_net = float(reward_amt or 0) - float(penalty_amt or 0)
        total_u = float(paid_u_dec or 0) + float(exp_u or 0) + exp_rmb_u + (rp_net if r.currency == "U" else rp_net / rate if rate else 0)
        sr = FinSalaryRecord(
            year_label=year_label,
            month=r.month,
            tg_name=r.tg_name,
            position=position,
            base_salary=base_salary,
            performance_salary=Decimal("0"),
            bonus=Decimal("0"),
            expense_rmb=exp_rmb if exp_rmb else None,
            expense_u=exp_u if exp_u else None,
            reward_amount=reward_amt if reward_amt else None,
            penalty_amount=penalty_amt if penalty_amt else None,
            paid_rmb=to_decimal(paid_rmb),
            paid_u=paid_u_dec,
            actual_amount=to_decimal(total_u) if total_u else None,
            exchange_rate=r.exchange_rate,
            wallet_address=wallet,
            period=_month_period(r.month) if r.month else None,
            remarks=r.remarks,
            payment_date=now_str()[:10],
            status="PAID",
            source="monthly_pay",
            created_at=now_str(), updated_at=now_str(),
        )
        db.add(sr)
        synced += 1

    db.commit()
    return ok(message=f"已付 {len(records)} 条，生成薪资记录 {synced} 条")


@router.delete("/api/finance/monthlypay/{rid}")
def delete_monthlypay(rid: int, db: Session = Depends(get_db)):
    db.query(FinMonthlyPayment).filter(FinMonthlyPayment.id == rid).delete()
    db.commit()
    return ok()


def _mp_dict(r):
    return {
        "id": r.id, "month": r.month, "team": r.team, "tgName": r.tg_name,
        "currency": r.currency,
        "amount": str(r.amount) if r.amount else None,
        "exchangeRate": str(r.exchange_rate) if r.exchange_rate else None,
        "remarks": r.remarks, "status": r.status,
    }


# ─── Excel 导入 ────────────────────────────────────────────

def _parse_month_from_period(period_str: str, year: str) -> str | None:
    """Extract YYYY-MM from period strings like '09.10～09.25', '11.01-11.30', '11月17日转正'."""
    import re
    if not period_str:
        return None
    s = str(period_str).strip()
    m = re.match(r'^(\d{1,2})[.\-～~]', s)
    if m:
        return f"{year}-{int(m.group(1)):02d}"
    m = re.match(r'^(\d{1,2})月', s)
    if m:
        return f"{year}-{int(m.group(1)):02d}"
    return None


def _extract_payment_date(row: tuple) -> str | None:
    """Try to extract a payment date string from a META row."""
    import re
    candidates = [str(v).strip() for v in row if v and isinstance(v, str)]
    for c in candidates:
        if re.search(r'月|日|支付|已发|发放', c):
            return c
    return None


@router.post("/api/finance/import/salary")
async def import_salary(file: UploadFile = File(...), year: str = Form(""), db: Session = Depends(get_db)):
    """Import salary records from standard Excel template.
    Columns: 序号/TG名称/岗位/基本工资/绩效工资RMB/奖金RMB/应发工资RMB/时间段/备注/收款地址TRC20/兑换比例/U数量/实际发放数量
    META rows (序号=None) may contain payment date info in col[10] or col[0]/col[1].
    """
    import re
    if not year:
        return err("请提供年份参数")
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.worksheets[0]

        # Two-pass: first collect all rows with their row index and type
        data_rows = []    # (row_idx, row_tuple)
        meta_rows = []    # (row_idx, payment_date_or_None)

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
            if not any(v for v in row):
                continue
            seq, tg = row[0], row[1]
            # Data row: seq is int AND tg_name is non-empty string
            if isinstance(seq, int) and tg and str(tg).strip():
                data_rows.append((i, row))
            else:
                # META row: may contain payment date
                pdate = _extract_payment_date(row)
                meta_rows.append((i, pdate))

        # Build a map: for each data_row index, find closest META row payment_date
        # Strategy: each data row gets the nearest META row payment_date
        #   - prefer the META row that comes immediately AFTER its batch (forward)
        #   - fall back to the META row just BEFORE (for pre-announce rows like row 30)
        meta_index = {idx: pdate for idx, pdate in meta_rows if pdate}

        def find_payment_date(data_idx: int) -> str | None:
            # Look for the nearest META row after this data row
            best_after = None
            best_before = None
            gap_after = float('inf')
            gap_before = float('inf')
            for midx, pdate in meta_index.items():
                if pdate:
                    if midx > data_idx and (midx - data_idx) < gap_after:
                        gap_after = midx - data_idx
                        best_after = pdate
                    elif midx < data_idx and (data_idx - midx) < gap_before:
                        gap_before = data_idx - midx
                        best_before = pdate
            # Prefer forward (batch summary), but only if within 15 rows
            if best_after and gap_after <= 15:
                return best_after
            if best_before and gap_before <= 5:
                return best_before
            return best_after or best_before

        def _col(r, i):
            return r[i] if i < len(r) else None

        imported = []
        for row_idx, row in data_rows:
            tg_name = str(row[1]).strip() if row[1] else None
            if not tg_name:
                continue
            period_str = str(_col(row, 7)).strip() if _col(row, 7) else None
            month = _parse_month_from_period(period_str, year)
            payment_date = find_payment_date(row_idx)

            sr = FinSalaryRecord(
                year_label=year,
                month=month,
                tg_name=tg_name,
                position=str(_col(row, 2)).strip() if _col(row, 2) else None,
                base_salary=to_decimal(_col(row, 3)),
                performance_salary=to_decimal(_col(row, 4)),
                bonus=to_decimal(_col(row, 5)),
                paid_rmb=to_decimal(_col(row, 6)),
                period=period_str,
                remarks=str(_col(row, 8)).strip() if _col(row, 8) else None,
                wallet_address=str(_col(row, 9)).strip() if _col(row, 9) else None,
                exchange_rate=to_decimal(_col(row, 10)),
                paid_u=to_decimal(_col(row, 11)),
                actual_amount=to_decimal(_col(row, 12)),
                payment_date=payment_date,
                status="PAID",
                source="excel",
                created_at=now_str(), updated_at=now_str(),
            )
            db.add(sr)
            imported.append({
                "month": month, "tg_name": tg_name,
                "position": str(row[2]).strip() if row[2] else "",
                "period": period_str,
                "paid_rmb": float(_col(row, 6)) if _col(row, 6) else None,
                "actual_amount": float(_col(row, 12)) if _col(row, 12) else None,
                "payment_date": payment_date or "",
            })

        db.commit()
        return ok(data=imported, message=f"导入完成：共 {len(imported)} 条薪资记录")
    except Exception as ex:
        db.rollback()
        import traceback
        return err(f"导入失败: {ex}\n{traceback.format_exc()}")


@router.post("/api/finance/import/fund")
async def import_fund(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        total = 0

        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                r = FinFundRecord(
                    period=ws.title,
                    network=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                    asset_type=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                    recharge_amount=to_decimal(row[5] if len(row) > 5 else None),
                    withdraw_amount=to_decimal(row[6] if len(row) > 6 else None),
                    contract_balance=to_decimal(row[0] if len(row) > 0 else None),
                    owner_balance=to_decimal(row[1] if len(row) > 1 else None),
                    record_date=str(row[2]).strip() if len(row) > 2 and row[2] else None,
                    remarks=str(row[7]).strip() if len(row) > 7 and row[7] else None,
                    created_at=now_str(), updated_at=now_str(),
                )
                db.add(r)
                total += 1

        db.commit()
        return ok(message=f"资金报表导入完成：{total} 条")
    except Exception as ex:
        db.rollback()
        return err(f"导入失败: {ex}")
